import pandas as pd
import seaborn as sns
import openpyxl
import re

book = openpyxl.load_workbook("processed_data_2.xlsx")
sheet = book.active
HH_VAC_LIST = ["bi","data","sql","аналитик данных","analyst","аналитик"]
sheet.cell(1,6).value = "Category"
for i in range(2, sheet.max_row+1):
  name = sheet.cell(i,1).value
  for reg in HH_VAC_LIST:
    regexp = re.compile(reg)
    if regexp.findall(str(name).lower()):
      sheet.cell(i,6).value = reg
      break

book.save("processed_data_3.xlsx")

df = pd.read_excel("processed_data.xlsx")

for city in df["Location"].unique():
  dfSeries = df[df["Location"]==city]
  for direct in dfSeries["Category"].unique():
    dfDirect = dfSeries[dfSeries["Category"]==direct]
    sns.histplot(dfDirect["Min_salary"])

TOP_5 = ["Москва", "Калининград","Ярославль","Санкт-Петербург","Набережные Челны"]
dfData = df[df["Category"]=="Аналитик данных"]
x = dfData[dfData.Location.isin(TOP_5)]
y = dfData["Max_salary"]
sns.boxplot(x=y,y=x["Location"], fill=False, width=.5, legend="brief")